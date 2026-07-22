"""Gold-set construction, the annotation spreadsheet round-trip, and extraction evaluation.

The gold set is ~600 human-verified reviews measuring extraction quality (facet P/R/F1 with
semantic matching, polarity accuracy over matched pairs). Two splits with different jobs:

- test (350): uniform random over the corpus -- uniform IS proportional, and this split's
  score is the datasheet claim. Touched exactly once, by the optimization winner.
- dev (250): stratified with the hard strata oversampled (trivial, long, foreign, low-star)
  -- reflection learns from failures, so failures are over-represented on purpose.

Hygiene: duplicate review texts removed before sampling; splits are product-disjoint (no
asin in both); fixed seed; membership recorded in a manifest. The annotation sheet carries
no star ratings (polarity-label anchoring), no model outputs (gold must be independent of
the system under test), and no split column (blind labeling; the manifest holds the map).

Aspect cell DSL: `facet: pos; facet: neg` -- or the literal `none` for "no aspects".
"""

from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path

import pandas as pd

TRIVIAL_MAX_LEN = 40
LONG_MIN_LEN = 400
US_COUNTRY = "the United States"

POLARITY_TOKENS = {
    "pos": "positive", "positive": "positive", "+": "positive",
    "neg": "negative", "negative": "negative", "-": "negative",
    "neu": "neutral", "neutral": "neutral", "0": "neutral",
}
REVIEWED_TOKENS = {"x", "ok", "yes", "done", "1", "true"}
FLAG_TOKENS = {"", "skip", "unsure"}

SHEET_COLUMNS = [
    "asin", "review_no", "stratum", "country", "product_title", "review_text",
    "draft_aspects", "gold_aspects", "reviewed", "flag", "rationale",
]

INSTRUCTIONS = """\
Gold-set annotation -- how to review

1. Read review_text (product_title gives context). Judge ONLY what the review discusses.
2. gold_aspects is pre-filled with the assistant's draft. Fix it in place:
   - format: facet: pos; facet: neg; facet: neu   (or exactly `none` for no aspects)
   - a facet is a short lowercase English noun phrase (1-4 words) naming a product
     DIMENSION: 'arch support', 'sizing accuracy', 'battery life'
   - polarity-neutral facets: the opinion goes in the polarity token, never the facet
   - product-intrinsic only: ignore delivery, seller, packaging condition, price paid,
     customer service
   - facets in English even when the review is not
3. The most likely draft error is OMISSION -- actively ask "is a facet missing?", not just
   "are these right?".
4. Mark reviewed with `x` when the row is done. Partial round-trips are fine.
5. flag = skip (garbage / not a review) or unsure (park for joint arbitration).
6. rationale: one short line on contested calls -- it feeds the prompt-optimization loop.
Never edit draft_aspects, asin, or review_no.
"""

_ILLEGAL_XLSX = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


# ------------------------------------------------------------------- sampling
def review_md5(text: str) -> str:
    return hashlib.md5(text.encode("utf-8")).hexdigest()


def assign_strata(df: pd.DataFrame) -> pd.Series:
    """First matching stratum wins: trivial, long, foreign, low_star, else typical."""
    text_len = df["text"].str.len()
    strata = {
        "trivial": text_len < TRIVIAL_MAX_LEN,
        "long": text_len > LONG_MIN_LEN,
        "foreign": df["country"].notna() & (df["country"] != US_COUNTRY),
        "low_star": pd.to_numeric(df["rev_stars"], errors="coerce") <= 2,
    }
    assigned = pd.Series("typical", index=df.index)
    taken = pd.Series(False, index=df.index)
    for name, mask in strata.items():
        pick = mask.fillna(False) & ~taken
        assigned[pick] = name
        taken |= pick
    return assigned


def make_gold_splits(
    reviews: pd.DataFrame, dev_n: int = 250, test_n: int = 350, seed: int = 0
) -> pd.DataFrame:
    """Sample the gold set. Returns rows with `split`, `stratum`, `review_md5`, shuffled.

    Test first (uniform over the deduplicated corpus), then dev from products absent from
    test (product-disjoint), stratified with even allocation over the five strata and any
    shortfall topped up from `typical`.
    """
    df = reviews.copy()
    df["text"] = df["text"].fillna("")
    df = df[df["text"].str.strip().str.len() > 0]
    df["review_md5"] = df["text"].map(review_md5)
    df = df.drop_duplicates("review_md5", keep="first")
    df["stratum"] = assign_strata(df)

    test = df.sample(test_n, random_state=seed).assign(split="test")

    pool = df[~df["asin"].isin(set(test["asin"]))]
    per_stratum = dev_n // 5
    parts = []
    for name in ["trivial", "long", "foreign", "low_star", "typical"]:
        sub = pool[pool["stratum"] == name]
        k = min(per_stratum, len(sub))
        if k:
            parts.append(sub.sample(k, random_state=seed))
    dev = pd.concat(parts)
    shortfall = dev_n - len(dev)
    if shortfall > 0:
        extra_pool = pool.loc[pool.index.difference(dev.index)]
        extra_pool = extra_pool[extra_pool["stratum"] == "typical"]
        dev = pd.concat([dev, extra_pool.sample(min(shortfall, len(extra_pool)), random_state=seed)])
    dev = dev.assign(split="dev")

    out = pd.concat([test, dev], ignore_index=True)
    return out.sample(frac=1, random_state=seed).reset_index(drop=True)


def write_manifest(splits: pd.DataFrame, path, seed: int) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    rows = splits[["asin", "review_no", "split", "stratum", "review_md5"]]
    payload = {"seed": seed, "n": len(rows), "rows": rows.to_dict(orient="records")}
    path.write_text(json.dumps(payload, indent=1))


def load_manifest(path) -> dict:
    """-> {(asin, review_no): {"split": ..., "stratum": ..., "review_md5": ...}}"""
    payload = json.loads(Path(path).read_text())
    return {(r["asin"], r["review_no"]): r for r in payload["rows"]}


# ------------------------------------------------------------------- aspects DSL
def format_aspects_dsl(aspects) -> str:
    """[(facet, polarity), ...] -> 'facet: pos; facet: neg' (or 'none')."""
    short = {"positive": "pos", "negative": "neg", "neutral": "neu"}
    if not aspects:
        return "none"
    return "; ".join(f"{f}: {short[p]}" for f, p in aspects)


def parse_aspects_dsl(cell: str | None):
    """Parse a gold_aspects cell. Returns (aspects, errors); aspects is None for a blank cell.

    Blank means "not filled in" (invalid on a reviewed row); the literal `none` means
    "explicitly no aspects". Facets are lowercased and whitespace-collapsed.
    """
    text = ("" if cell is None else str(cell)).strip()
    if text == "":
        return None, []
    if text.lower() in {"none", "-"}:
        return [], []
    aspects, errors = [], []
    for item in filter(None, (p.strip() for p in text.split(";"))):
        if ":" not in item:
            errors.append(f"missing ':' in {item!r}")
            continue
        facet, _, pol = item.rpartition(":")
        facet = re.sub(r"\s+", " ", facet.strip().lower())
        pol = pol.strip().lower()
        if not facet:
            errors.append(f"empty facet in {item!r}")
        elif pol not in POLARITY_TOKENS:
            errors.append(f"bad polarity {pol!r} in {item!r} (use pos/neg/neu)")
        else:
            aspects.append((facet, POLARITY_TOKENS[pol]))
    return aspects, errors


# ------------------------------------------------------------------- xlsx round-trip
def _clean(value) -> str:
    return _ILLEGAL_XLSX.sub("", "" if value is None else str(value))


def write_annotation_sheet(rows: pd.DataFrame, path) -> None:
    """Write the annotation workbook: a `gold` sheet plus an `instructions` sheet.

    `rows` needs SHEET_COLUMNS minus the annotator columns (draft/gold/reviewed/flag/
    rationale), which start empty and are filled by merge_drafts / the reviewer.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "gold"
    widths = [14, 10, 10, 12, 40, 80, 40, 40, 9, 8, 40]
    ws.append(SHEET_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wrap = Alignment(wrap_text=True, vertical="top")
    grey = PatternFill("solid", fgColor="EEEEEE")
    for row in rows.itertuples():
        ws.append([
            _clean(row.asin), int(row.review_no), _clean(row.stratum), _clean(row.country),
            _clean(row.product_title), _clean(row.text), "", "", "", "", "",
        ])
        r = ws.max_row
        for col in (5, 6, 7, 8, 11):
            ws.cell(row=r, column=col).alignment = wrap
        ws.cell(row=r, column=7).fill = grey  # draft column: visually read-only

    n = len(rows) + 1
    dv_rev = DataValidation(type="list", formula1='"x"', allow_blank=True)
    dv_flag = DataValidation(type="list", formula1='"skip,unsure"', allow_blank=True)
    ws.add_data_validation(dv_rev)
    ws.add_data_validation(dv_flag)
    dv_rev.add(f"I2:I{n}")
    dv_flag.add(f"J2:J{n}")
    ws.freeze_panes = "A2"

    info = wb.create_sheet("instructions")
    info.column_dimensions["A"].width = 100
    for line in INSTRUCTIONS.splitlines():
        info.append([line])
    wb.save(path)


def merge_drafts(sheet_path, drafts) -> int:
    """Inject drafts into the sheet: sets draft_aspects and pre-fills gold_aspects.

    `drafts` is an iterable of {"asin", "review_no", "draft"} where draft is DSL text.
    gold_aspects is only pre-filled when still empty (never clobbers reviewer edits).
    Returns the number of rows updated.
    """
    from openpyxl import load_workbook

    wb = load_workbook(sheet_path)
    ws = wb["gold"]
    index = {
        (ws.cell(row=r, column=1).value, int(ws.cell(row=r, column=2).value)): r
        for r in range(2, ws.max_row + 1)
    }
    updated = 0
    for d in drafts:
        r = index.get((d["asin"], int(d["review_no"])))
        if r is None:
            continue
        ws.cell(row=r, column=7, value=d["draft"])
        if not ws.cell(row=r, column=8).value:
            ws.cell(row=r, column=8, value=d["draft"])
        updated += 1
    wb.save(sheet_path)
    return updated


def ingest_reviewed_sheet(sheet_path, manifest_path, out_dir) -> dict:
    """Validate the reviewed sheet and freeze gold_dev.jsonl / gold_test.jsonl.

    Only rows marked reviewed and not flagged are frozen; `unsure` rows are returned for
    joint arbitration; `skip` rows are recorded. Returns a report with per-row errors,
    counts, and draft-vs-gold agreement stats (label provenance for the datasheet).
    """
    manifest = load_manifest(manifest_path)
    sheet = pd.read_excel(sheet_path, sheet_name="gold", dtype=str).fillna("")
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    report = {
        "total": len(sheet), "reviewed": 0, "unreviewed": 0, "skipped": 0,
        "unsure": [], "errors": [], "frozen": {"dev": 0, "test": 0},
        "agreement": {"rows_compared": 0, "rows_unchanged": 0,
                      "facets_added": 0, "facets_removed": 0, "polarity_changed": 0},
    }
    frozen = {"dev": [], "test": []}

    for i, row in sheet.iterrows():
        excel_row = i + 2
        key = (row["asin"], int(row["review_no"]))
        if key not in manifest:
            report["errors"].append(f"row {excel_row}: {key} not in manifest")
            continue
        flag = row["flag"].strip().lower()
        if flag not in FLAG_TOKENS:
            report["errors"].append(f"row {excel_row}: bad flag {row['flag']!r}")
            continue
        if flag == "skip":
            report["skipped"] += 1
            continue
        if flag == "unsure":
            report["unsure"].append({"row": excel_row, "asin": key[0], "review_no": key[1],
                                     "rationale": row["rationale"]})
            continue
        if row["reviewed"].strip().lower() not in REVIEWED_TOKENS:
            report["unreviewed"] += 1
            continue

        gold, errors = parse_aspects_dsl(row["gold_aspects"])
        if gold is None:
            errors = ["gold_aspects is blank on a reviewed row (use `none` for no aspects)"]
        if errors:
            report["errors"].extend(f"row {excel_row}: {e}" for e in errors)
            continue

        report["reviewed"] += 1
        draft, draft_errors = parse_aspects_dsl(row["draft_aspects"])
        if draft is not None and not draft_errors:
            agr = report["agreement"]
            agr["rows_compared"] += 1
            draft_set, gold_set = dict(draft), dict(gold)
            agr["facets_added"] += len(set(gold_set) - set(draft_set))
            agr["facets_removed"] += len(set(draft_set) - set(gold_set))
            agr["polarity_changed"] += sum(
                1 for f in set(draft_set) & set(gold_set) if draft_set[f] != gold_set[f]
            )
            agr["rows_unchanged"] += int(sorted(draft) == sorted(gold))

        split = manifest[key]["split"]
        frozen[split].append({
            "asin": key[0], "review_no": key[1], "text": row["review_text"],
            "stratum": manifest[key]["stratum"],
            "gold_aspects": [{"facet": f, "polarity": p} for f, p in gold],
            "rationale": row["rationale"],
        })

    for split, records in frozen.items():
        out = out_dir / f"gold_{split}.jsonl"
        with open(out, "w") as fh:
            for rec in records:
                fh.write(json.dumps(rec) + "\n")
        report["frozen"][split] = len(records)
    return report


# ------------------------------------------------------------------- evaluation
def load_gold(path) -> dict:
    """Load a frozen gold file -> {(asin, review_no): [(facet, polarity), ...]}."""
    gold = {}
    with open(path) as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            rec = json.loads(line)
            gold[(rec["asin"], rec["review_no"])] = [
                (a["facet"].strip().lower(), a["polarity"]) for a in rec["gold_aspects"]
            ]
    return gold


def exact_match(pred_facet: str, gold_facet: str) -> bool:
    return pred_facet == gold_facet


def cosine_matcher(threshold: float = 0.80):
    """Semantic facet matcher on the shared dense encoder ("sizing" ~ "sizing accuracy").

    Runs on CPU deliberately: under memory pressure (e.g. a large Ollama model resident),
    MPS command buffers can fail with GPU-OOM and silently return garbage embeddings --
    identity pairs stop matching and every score deflates. CPU is immune, and this model
    is small enough that matcher throughput is irrelevant. A canary identity check guards
    against any other silent-corruption mode.
    """
    from sentence_transformers import SentenceTransformer

    from emmr import config

    model = SentenceTransformer(config.DENSE_ENCODER, device="cpu")
    cache: dict = {}

    def embed(facet: str):
        if facet not in cache:
            cache[facet] = model.encode(facet, normalize_embeddings=True)
        return cache[facet]

    def match(pred_facet: str, gold_facet: str) -> bool:
        return float(embed(pred_facet) @ embed(gold_facet)) >= threshold

    if not match("canary facet", "canary facet"):  # identity must match at any threshold <= 1
        raise RuntimeError("cosine_matcher self-check failed: encoder returned corrupt embeddings")
    return match


def evaluate_extraction(pred: dict, gold: dict, match=exact_match) -> dict:
    """Facet P/R/F1 and polarity accuracy of predictions against the gold labels.

    `pred` and `gold` map (asin, review_no) -> list of (facet, polarity). Matching is greedy
    one-to-one per review using `match` on the facet strings; polarity accuracy is computed
    over matched pairs only. Reviews present in gold but absent from pred count all their
    gold facets as misses.
    """
    tp = fp = fn = 0
    polarity_hits = polarity_total = 0

    for key, gold_aspects in gold.items():
        pred_aspects = list(pred.get(key, []))
        remaining = list(gold_aspects)
        for p_facet, p_polarity in pred_aspects:
            hit = next((g for g in remaining if match(p_facet, g[0])), None)
            if hit is None:
                fp += 1
            else:
                remaining.remove(hit)
                tp += 1
                polarity_total += 1
                polarity_hits += int(p_polarity == hit[1])
        fn += len(remaining)

    precision = tp / (tp + fp) if tp + fp else 0.0
    recall = tp / (tp + fn) if tp + fn else 0.0
    f1 = 2 * precision * recall / (precision + recall) if precision + recall else 0.0
    return {
        "facet_precision": round(precision, 4),
        "facet_recall": round(recall, 4),
        "facet_f1": round(f1, 4),
        "polarity_accuracy": round(polarity_hits / polarity_total, 4) if polarity_total else None,
        "n_reviews": len(gold),
    }

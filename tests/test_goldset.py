import json

import pandas as pd
import pytest

from emmr.reviews.goldset import (
    format_aspects_dsl,
    ingest_reviewed_sheet,
    load_gold,
    make_gold_splits,
    merge_drafts,
    parse_aspects_dsl,
    write_annotation_sheet,
    write_manifest,
)


# ---------------------------------------------------------------------- DSL
def test_dsl_round_trip():
    aspects = [("arch support", "positive"), ("sizing accuracy", "negative")]
    cell = format_aspects_dsl(aspects)
    assert cell == "arch support: pos; sizing accuracy: neg"
    parsed, errors = parse_aspects_dsl(cell)
    assert parsed == aspects and errors == []


def test_dsl_none_vs_blank():
    assert parse_aspects_dsl("none") == ([], [])
    assert format_aspects_dsl([]) == "none"
    parsed, errors = parse_aspects_dsl("   ")
    assert parsed is None and errors == []  # blank = not filled in, distinct from none


def test_dsl_normalises_and_reports_errors():
    parsed, errors = parse_aspects_dsl("  Arch   Support : POS; waterproofing: wrong; : neg")
    assert parsed == [("arch support", "positive")]
    assert len(errors) == 2  # bad polarity token + empty facet


# ---------------------------------------------------------------------- splits
def _reviews(n=400, seed=1):
    import numpy as np

    rng = np.random.default_rng(seed)
    return pd.DataFrame({
        "asin": [f"A{i % 200}" for i in range(n)],
        "review_no": [i // 200 for i in range(n)],
        "text": [
            f"ok {i}" if i % 17 == 0                                # trivial (<40 chars), unique
            else (f"long review {i} " + "padding " * 60 if i % 13 == 0   # long (>400), unique
                  else f"a decently sized review number {i} with enough words to be typical text")
            for i in range(n)
        ],
        "rev_stars": rng.choice([1, 2, 3, 4, 5], size=n).astype(float),
        "country": ["Australia" if i % 11 == 0 else "the United States" for i in range(n)],
    })


def test_make_gold_splits_sizes_and_product_disjointness():
    splits = make_gold_splits(_reviews(), dev_n=50, test_n=60, seed=0)
    assert (splits.split == "test").sum() == 60
    assert (splits.split == "dev").sum() == 50
    dev_asins = set(splits[splits.split == "dev"].asin)
    test_asins = set(splits[splits.split == "test"].asin)
    assert not dev_asins & test_asins  # product-disjoint
    assert splits.review_md5.is_unique  # text-level dedup


def test_make_gold_splits_dev_oversamples_hard_strata():
    splits = make_gold_splits(_reviews(), dev_n=50, test_n=60, seed=0)
    dev_strata = splits[splits.split == "dev"].stratum.value_counts()
    assert {"trivial", "long", "foreign", "low_star"} <= set(dev_strata.index)


def test_make_gold_splits_reproducible():
    a = make_gold_splits(_reviews(), dev_n=30, test_n=40, seed=7)
    b = make_gold_splits(_reviews(), dev_n=30, test_n=40, seed=7)
    assert a[["asin", "review_no", "split"]].equals(b[["asin", "review_no", "split"]])


# ---------------------------------------------------------------------- xlsx round trip
@pytest.fixture
def annotation_setup(tmp_path):
    splits = make_gold_splits(_reviews(), dev_n=10, test_n=10, seed=0)
    splits["product_title"] = "Some Product"
    sheet = tmp_path / "gold_annotation.xlsx"
    manifest = tmp_path / "manifest.json"
    write_manifest(splits, manifest, seed=0)
    write_annotation_sheet(splits, sheet)
    return splits, sheet, manifest, tmp_path


def test_sheet_write_merge_ingest_round_trip(annotation_setup):
    splits, sheet, manifest, tmp_path = annotation_setup
    first = splits.iloc[0]

    drafts = [{"asin": first.asin, "review_no": int(first.review_no),
               "draft": "comfort: pos; sizing accuracy: neg"}]
    assert merge_drafts(sheet, drafts) == 1

    # simulate the reviewer: fix one polarity, add a facet, mark reviewed
    from openpyxl import load_workbook

    wb = load_workbook(sheet)
    ws = wb["gold"]
    row = next(r for r in range(2, ws.max_row + 1)
               if ws.cell(r, 1).value == first.asin
               and int(ws.cell(r, 2).value) == int(first.review_no))
    ws.cell(row, 8, value="comfort: neg; sizing accuracy: neg; lace durability: neg")
    ws.cell(row, 9, value="x")
    ws.cell(row, 11, value="comfort was actually criticised")
    wb.save(sheet)

    report = ingest_reviewed_sheet(sheet, manifest, tmp_path)

    assert report["reviewed"] == 1
    assert report["unreviewed"] == len(splits) - 1
    assert report["errors"] == []
    agr = report["agreement"]
    assert agr["rows_compared"] == 1 and agr["rows_unchanged"] == 0
    assert agr["facets_added"] == 1      # lace durability
    assert agr["polarity_changed"] == 1  # comfort pos -> neg

    split = json.loads((manifest).read_text())
    split_of = {(r["asin"], r["review_no"]): r["split"] for r in split["rows"]}
    frozen = load_gold(tmp_path / f"gold_{split_of[(first.asin, int(first.review_no))]}.jsonl")
    assert frozen[(first.asin, int(first.review_no))] == [
        ("comfort", "negative"), ("sizing accuracy", "negative"), ("lace durability", "negative"),
    ]


def test_ingest_validation_catches_blank_and_bad_rows(annotation_setup):
    splits, sheet, manifest, tmp_path = annotation_setup
    from openpyxl import load_workbook

    wb = load_workbook(sheet)
    ws = wb["gold"]
    ws.cell(2, 9, value="x")                      # reviewed but gold_aspects blank
    ws.cell(3, 8, value="comfort: maybe")
    ws.cell(3, 9, value="x")                      # bad polarity token
    ws.cell(4, 10, value="unsure")                # parked for arbitration
    wb.save(sheet)

    report = ingest_reviewed_sheet(sheet, manifest, tmp_path)
    assert len(report["errors"]) == 2
    assert any("blank" in e for e in report["errors"])
    assert any("bad polarity" in e for e in report["errors"])
    assert len(report["unsure"]) == 1
